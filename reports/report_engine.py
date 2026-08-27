from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_BG = "1A3A5C"
HEADER_FG = "FFFFFF"
ALT_ROW_BG = "F0F4F8"

DENOM_MAP = {
    "€5":    "5 EUR",
    "€10":   "10 EUR",
    "€20":   "20 EUR",
    "€50":   "50 EUR",
    "€100":  "100 EUR",
    "€200":  "200 EUR",
    "€500":  "500 EUR",
    "Coins": "Bulk Coin",
}

COLUMNS = [
    "Collection Date",
    "Processed Date",
    "Customer ID",
    "Customer Name",
    "Customer Location",
    "Address1",
    "Seal Number",
    "Wallet Number",
    "Lodgement Slip Number",
    "Operator Id",
    "Comments",
    "Expected Total",
    "Discrepancy",
    "5 EUR",
    "10 EUR",
    "20 EUR",
    "50 EUR",
    "100 EUR",
    "200 EUR",
    "500 EUR",
    "Bulk Coin",
    "Coin Processed",
    "Notes Processed",
    "Total Processed",
    "Payment",
    "Adjustment Reason",
]

COL_WIDTHS = {
    "Collection Date": 16,
    "Processed Date": 16,
    "Customer ID": 14,
    "Customer Name": 22,
    "Customer Location": 22,
    "Address1": 20,
    "Seal Number": 16,
    "Wallet Number": 16,
    "Lodgement Slip Number": 22,
    "Operator Id": 14,
    "Comments": 20,
    "Expected Total": 16,
    "Discrepancy": 14,
    "5 EUR": 10,
    "10 EUR": 10,
    "20 EUR": 10,
    "50 EUR": 10,
    "100 EUR": 10,
    "200 EUR": 10,
    "500 EUR": 10,
    "Bulk Coin": 12,
    "Coin Processed": 16,
    "Notes Processed": 16,
    "Total Processed": 16,
    "Payment": 14,
    "Adjustment Reason": 22,
}


def _thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# S-07: CSV/Excel formula injection. Any string value starting with one of
# these characters is interpreted as a formula by Excel/LibreOffice/Sheets
# when the exported file is opened -- applied generically to every string
# cell this report writes (not just wallet_id, the field the review called
# out), since any user-derived field could carry the same risk if the data
# model ever grows a new one.
_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@")


def _neutralize_formula_injection(value):
    if isinstance(value, str) and value.startswith(_FORMULA_TRIGGER_CHARS):
        return "'" + value
    return value


def _build_row(t: dict) -> dict:
    denoms = {d["denomination"]: d for d in t.get("denominations", [])}

    def count(label):
        return denoms[label]["count"] if label in denoms else ""

    def val(label):
        return float(denoms[label]["value"]) if label in denoms else 0.0

    coin_value = val("Coins")
    notes_value = sum(
        val(k) for k in ("€5", "€10", "€20", "€50", "€100", "€200", "€500")
    )
    total = float(t.get("total_value", 0))
    raw_expected = t.get("expected_total")
    expected = float(raw_expected) if raw_expected is not None else None
    discrepancy = round(total - expected, 2) if expected is not None else ""
    date_str = (
        t["created_at"].strftime("%d/%m/%Y")
        if isinstance(t["created_at"], datetime)
        else str(t.get("created_at", ""))
    )

    row = {
        "Collection Date":       date_str,
        "Processed Date":        date_str,
        "Customer ID":           t.get("customer_id", ""),
        "Customer Name":         t.get("customer_name", ""),
        "Customer Location":     t.get("location_name", ""),
        "Address1":              "",
        "Seal Number":           t.get("bag_number", ""),
        "Wallet Number":         t.get("wallet_id") or "",
        "Lodgement Slip Number": "",
        "Operator Id":           t.get("username", ""),
        "Comments":              "",
        "Expected Total":        expected if expected is not None else total,
        "Discrepancy":           discrepancy,
        "5 EUR":                 count("€5"),
        "10 EUR":                count("€10"),
        "20 EUR":                count("€20"),
        "50 EUR":                count("€50"),
        "100 EUR":               count("€100"),
        "200 EUR":               count("€200"),
        "500 EUR":               count("€500"),
        "Bulk Coin":             count("Coins"),
        "Coin Processed":        coin_value if coin_value else "",
        "Notes Processed":       round(notes_value, 2) if notes_value else "",
        "Total Processed":       total,
        "Payment":               "",
        "Adjustment Reason":     "",
    }
    # S-07: neutralize formula injection generically across every column,
    # applied once here so both the xlsx and csv writers below get it for
    # free without needing their own sanitization pass.
    return {key: _neutralize_formula_injection(val) for key, val in row.items()}


def build_transactions_excel(transactions: list[dict], output_path: str) -> str:
    rows = [_build_row(t) for t in transactions]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    ncols = len(COLUMNS)

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    title_cell = ws["A1"]
    title_cell.value = "BRINK'S NEXUS — Cash Processing Report"
    title_cell.font = Font(bold=True, size=13, color=HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Timestamp row
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    ts_cell = ws["A2"]
    ts_cell.value = (
        f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}  |  "
        f"Total records: {len(rows)}"
    )
    ts_cell.font = Font(italic=True, size=9, color="666666")
    ts_cell.alignment = Alignment(horizontal="right")
    ws.row_dimensions[2].height = 14

    # Header row
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    header_font = Font(bold=True, color=HEADER_FG, size=9)
    for col_idx, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=4, column=col_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col, 14)
    ws.row_dimensions[4].height = 30

    border = _thin_border()
    alt_fill = PatternFill("solid", fgColor=ALT_ROW_BG)
    currency_cols = {"Expected Total", "Coin Processed", "Notes Processed", "Total Processed"}
    numeric_cols = {"5 EUR", "10 EUR", "20 EUR", "50 EUR", "100 EUR", "200 EUR", "500 EUR", "Bulk Coin"}

    for row_idx, row in enumerate(rows, 5):
        is_alt = (row_idx % 2 == 0)
        for col_idx, col in enumerate(COLUMNS, 1):
            val = row[col]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if is_alt:
                cell.fill = alt_fill
            if col in currency_cols and isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col in numeric_cols and isinstance(val, int):
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(ncols)}4"

    wb.save(output_path)
    return output_path


def build_transactions_csv(transactions: list[dict], output_path: str) -> str:
    rows = [_build_row(t) for t in transactions]
    df = pd.DataFrame(rows, columns=COLUMNS)
    df.to_csv(output_path, index=False)
    return output_path
